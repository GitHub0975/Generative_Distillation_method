#!/usr/bin/env python
# coding: utf-8

# # DEMO ON MNIST SPLIT


from MAS import *
from CIFAR_split import *
from iCIFAR100 import iCIFAR100
import shutil
import matplotlib.ticker as ticker
import torch.nn as nn

data_path = 'D:\\Distillation/'
batch_size=100
kwargs = {'num_workers': 0, 'pin_memory': True} 


        
base_class_size = 20
inc_class_size = 20
total_class = 100

exp_dirname = 'Result_vgg' + str(base_class_size) + '_' + str(inc_class_size)
data_dirname = 'cifar' + str(base_class_size) + '_' + str(inc_class_size)



# 將MNIST dataset切成五個任務的資料儲存(放進loader就可以用了)
# 建立TASK1~TASK50
class_split = [[i for i in range(base_class_size)]]
# 5類算一個task
for i in range(base_class_size, total_class, inc_class_size):
    #print(np.arange(i,i+10))
    class_split.append([j for j in range(i, i+inc_class_size)])
print(class_split)
#exit()

transform_train_resnet = transforms.Compose([
                           transforms.RandomCrop(32, padding=4),
                           transforms.RandomHorizontalFlip(),
                           transforms.ToTensor(),
                           #transforms.Normalize(mean = [0.485, 0.456, 0.406], std = [0.229, 0.224, 0.225])
                           transforms.Normalize(mean=[0.507, 0.487, 0.441], std=[0.267, 0.256, 0.276])
                       ])
                       
tranform_test_resnet = transforms.Compose([
                           transforms.ToTensor(),
                           #transforms.Normalize(mean = [0.485, 0.456, 0.406], std = [0.229, 0.224, 0.225])
                           transforms.Normalize(mean=[0.507, 0.487, 0.441], std=[0.267, 0.256, 0.276])
                       ])
                       
transform_train_vgg = transforms.Compose([
                           transforms.RandomCrop((32,32),padding=4),
                           transforms.RandomHorizontalFlip(p=0.5),
                           transforms.ColorJitter(brightness=0.24705882352941178),  
                           transforms.ToTensor(),
                           transforms.Normalize(mean = [0.485, 0.456, 0.406], std = [0.229, 0.224, 0.225])
                           #transforms.Normalize(mean=[0.507, 0.487, 0.441], std=[0.267, 0.256, 0.276])
                       ])
                       
transform_test_vgg = transforms.Compose([
                           transforms.ToTensor(),
                           transforms.Normalize(mean = [0.485, 0.456, 0.406], std = [0.229, 0.224, 0.225])
                           #transforms.Normalize(mean=[0.507, 0.487, 0.441], std=[0.267, 0.256, 0.276])
                       ])



task = 1
for digits in class_split:
    #print(digits)                
    # 使用label的真值(0-9作為label)
    dsets = {}
    #dsets['train_true']=    CIFAR_Split(data_path, train=True, download=True,
    #                   transform=transform_train_vgg,digits=digits)
    classes = [digits[0], digits[-1] + 1]

    dsets['train_true'] = iCIFAR100(data_path, transform=transform_train_vgg, download=True, classes=classes)
    
    dsets['val_true'] = iCIFAR100('../dataset', test_transform=transform_test_vgg, train=False, download=True, classes = classes)
    dlabel=str(task)
    
    if not os.path.isdir('data/' + data_dirname):
        os.mkdir('data/' + data_dirname)
    torch.save(dsets,'data/' + data_dirname +'/split'+dlabel+'_dataset.pth.tar')
    task += 1
#exit()

task = 1
if not os.path.isdir('./General_utils/information'):
    os.mkdir('./General_utils/information')

#FIRST TASK TRAINING
from MNIST_Net import *
test = Fearnet(num_classes = base_class_size)
#test.apply(init_weights)
torch.save(test, 'General_utils/Fearnet' + str(base_class_size) + '_' + str(inc_class_size)+ '.pth.tar')

model_path='General_utils/Fearnet' + str(base_class_size) + '_' + str(inc_class_size)+ '.pth.tar'
from Finetune_SGD import *
#digits = [1,2]
digits = class_split[0]
dlabel=str(task)


dataset_path=data_path + 'data/' + data_dirname + '/split'+dlabel+'_dataset.pth.tar'

exp_dir='exp_dir/' + exp_dirname + '/cifar_NET_normalize'+dlabel

num_epochs=250

fine_tune_SGD(dataset_path=dataset_path, num_epochs=num_epochs,
    exp_dir=exp_dir,model_path=model_path,lr=1e-3,batch_size=128, class_num=len(digits))
#model_path=os.path.join(exp_dir,'best_model.pth.tar')
#exit()

    

#MIMIC the case when samples from the previous takss are seen in each step
from MAS import *

reg_sets=[]
dataset_path = data_path + 'data/' + data_dirname + '//split1_dataset.pth.tar'

#exp_dir='exp_dir/Test/cifar_NET_normalize1'
exp_dir='exp_dir/' + exp_dirname + '/cifar_NET_normalize1'
previous_pathes=[]
reg_lambda=10

history_task_acc=np.zeros((51, 51))
history_task_acc[0][0] = 0.71    # 之後要再改寫
total_mean_acc = np.zeros((51, ))
total_mean_acc[0] = 0.71
#task = 7
for digits in class_split[1:]:
    reg_sets.append(dataset_path)    # 舊資料

    model_path=os.path.join(exp_dir,'epoch.pth.tar')
    previous_pathes.append(model_path)    # 每次的新任務都加上上一次訓練好的模型
    print(model_path)
    
    task += 1
    #task = 50
    dlabel=str(task)

    dataset_path = data_path + 'data/' + data_dirname + '//split'+dlabel+'_dataset.pth.tar'    # 新任務的資料路徑

    exp_dir='exp_dir/' + exp_dirname+'/cifar_NET_normalize'+dlabel
   
    num_epochs=250
    
    data_dirs=None
    
    dsets = torch.load(dataset_path)

    class_num = digits[-1] + 1     # 目前有幾個類別

    print('Total training class number: {}'.format(class_num))
    #mean_acc = 0
    #task_acc = [0]
    #if digits[0] >10:
    model_ft, task_acc, mean_acc = Big_Separate(dataset_path=dataset_path, previous_pathes=previous_pathes, 
                previous_task_model_path=model_path, exp_dir=exp_dir, data_dirs=data_dirs, 
                reg_sets=reg_sets, reg_lambda=reg_lambda, batch_size=128, num_epochs=num_epochs, lr=1e-3,norm='L2', 
                b1=False, class_num = class_num, task = task, digits = digits, base_class_size=base_class_size, inc_class_size=inc_class_size)
   
    # 紀錄每個task之後的mean_acc
    total_mean_acc[task-1] = mean_acc
    
    # 紀錄每個task的歷史紀錄
    for i in range(len(task_acc)):
        history_task_acc[i][task-1] = task_acc[i]
    
    print("Final Model's state_dict:")
    #print(model_ft)
    layer_names = []
    #for param_tensor in model_ft.state_dict():
    #    layer_names.append(param_tensor)
    #    print(param_tensor, "\t", model_ft.state_dict()[param_tensor].size())

print('total_mean_acc')   
print(total_mean_acc) 
    
    
# 儲存歷史記錄圖
fig, ax = plt.subplots(2,2)
plt.xticks(fontsize=20)

fig.set_figheight(25)
fig.set_figwidth(70)

        
colormap = plt.cm.gist_ncar
colors = [colormap(i) for i in np.random.rand(12)]

for i in range(12):
    legend_label = 'task' + str(i)
    ax[0, 0].plot(history_task_acc[i], 'o--', color = colors[i], label = legend_label)
ax[0,0].legend(bbox_to_anchor=(0.5, 1.15), ncol=3)
ax[0,0].set_xlabel('number of tasks')
ax[0,0].axis(xmin=0, xmax=51)
ax[0,0].axis(ymin=0, ymax=1)

ax[0,0].xaxis.set_major_locator(ticker.MultipleLocator(1))
ax[0,0].yaxis.set_major_locator(ticker.MultipleLocator(0.1))
ax[0,0].grid(True)

colors = [colormap(i) for i in np.random.rand(12)]
for i in range(12, 24):
    legend_label = 'task' + str(i)
    ax[0, 1].plot(history_task_acc[i], 'o--', color = colors[i-12], label = legend_label)
ax[0,1].legend(bbox_to_anchor=(0.5, 1.15), ncol=3)
ax[0,1].set_xlabel('number of tasks')
ax[0,1].axis(xmin=0, xmax=51)
ax[0,1].axis(ymin=0, ymax=1)

ax[0,1].xaxis.set_major_locator(ticker.MultipleLocator(1))
ax[0,1].yaxis.set_major_locator(ticker.MultipleLocator(0.1))
ax[0,1].grid(True)

colors = [colormap(i) for i in np.random.rand(12)]
for i in range(24, 36):
    legend_label = 'task' + str(i)
    ax[1, 0].plot(history_task_acc[i], 'o--', color = colors[i-24], label = legend_label)
ax[1,0].legend(bbox_to_anchor=(0.5, 1.15), ncol=3)
ax[1,0].set_xlabel('number of tasks')
ax[1,0].axis(xmin=0, xmax=51)
ax[1,0].axis(ymin=0, ymax=1)

ax[1,0].xaxis.set_major_locator(ticker.MultipleLocator(1))
ax[1,0].yaxis.set_major_locator(ticker.MultipleLocator(0.1))
ax[1,0].grid(True)

colors = [colormap(i) for i in np.random.rand(15)]
for i in range(36, len(history_task_acc)):
    legend_label = 'task' + str(i)
    ax[1, 1].plot(history_task_acc[i], 'o--', color = colors[i-36], label = legend_label)
ax[1,1].legend(bbox_to_anchor=(0.5, 1.15), ncol=3)
ax[1,1].set_xlabel('number of tasks')
ax[1,1].axis(xmin=0, xmax=51)
ax[1,1].axis(ymin=0, ymax=1)

ax[1,1].xaxis.set_major_locator(ticker.MultipleLocator(1))
ax[1,1].yaxis.set_major_locator(ticker.MultipleLocator(0.1))
ax[1,1].grid(True)


# In[24]:
# Avg_acc
Avg_acc = []
for i in range(task):
    Avg_acc.append(np.sum(np.array(history_task_acc)[:,i]) / (i + 1))
    #print(Avg_acc)
plt.figure(figsize=(70,25))
plt.plot(Avg_acc, 'o--')
plt.axis(xmin=0, xmax=51)
plt.grid()
plt.xticks(np.arange(0, 50, 1))
plt.yticks(np.arange(0, 1, 0.1))
    #plt.xlim(50, 100)
#plt.show()


print(task_acc)

import openpyxl
wb = openpyxl.load_workbook('Acc_result.xlsx')
sheet = wb['工作表1']
c = sheet.cell(row = 1, column = 1)
j = 1
while c.value is not None:
    j += 1
    c = sheet.cell(row=1, column = j)
#j += 1
for i in range(len(task_acc)):
    sheet.cell(row = i+2, column = j, value = task_acc[i])

wb.create_sheet('history_acc')
sh = wb['history_acc']

for i in range(51):
    sh.cell(row = 1, column = i+1, value = 'task' + str(i+1))

for i in range(51):
    for k in range(51):
        sh.cell(row = i+2, column = k+1, value = history_task_acc[i][k])

# 儲存mean_acc        
sheet = wb['total_mean_acc']
c = sheet.cell(row = 1, column = 1)
j = 1
while c.value is not None:
    j += 1
    c = sheet.cell(row=1, column = j)
#j += 1
for i in range(len(total_mean_acc)):
    sheet.cell(row = i+2, column = j, value = total_mean_acc[i])

    
wb.save('Acc_result.xlsx')

